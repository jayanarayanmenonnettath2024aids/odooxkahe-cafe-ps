"""
Reports router — dashboard, metrics, export.
"""

import io
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.dependencies import AdminUser
from app.schemas.common import SuccessResponse
from app.schemas.report import DashboardResponse, ReportFilters
from app.services.report_service import ReportService

router = APIRouter(prefix="/reports", tags=["Reports"])


@router.get("/dashboard", response_model=SuccessResponse[DashboardResponse])
async def get_dashboard(
    period: Optional[str] = Query(None, description="today, week, month, custom"),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    employee_id: Optional[int] = Query(None),
    session_id: Optional[int] = Query(None),
    product_id: Optional[int] = Query(None),
    admin = Depends(AdminUser),
    db: AsyncSession = Depends(get_db),
):
    filters = ReportFilters(
        period=period,
        start_date=start_date,
        end_date=end_date,
        employee_id=employee_id,
        session_id=session_id,
        product_id=product_id,
    )
    service = ReportService(db)
    dashboard = await service.get_dashboard(filters)
    return SuccessResponse(data=dashboard)


@router.get("/export")
async def export_report(
    format: str = Query("xlsx", description="pdf or xlsx"),
    period: Optional[str] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    employee_id: Optional[int] = Query(None),
    session_id: Optional[int] = Query(None),
    admin = Depends(AdminUser),
    db: AsyncSession = Depends(get_db),
):
    """Export reports as PDF or XLSX."""
    filters = ReportFilters(
        period=period,
        start_date=start_date,
        end_date=end_date,
        employee_id=employee_id,
        session_id=session_id,
    )
    service = ReportService(db)
    dashboard = await service.get_dashboard(filters)

    if format == "xlsx":
        import openpyxl
        wb = openpyxl.Workbook()

        # Summary sheet
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Metric", "Value"])
        ws.append(["Total Orders", dashboard.total_orders])
        ws.append(["Total Revenue", dashboard.total_revenue])
        ws.append(["Average Order Value", dashboard.average_order_value])

        # Top Products sheet
        ws2 = wb.create_sheet("Top Products")
        ws2.append(["Product", "Quantity Sold", "Revenue"])
        for p in dashboard.top_products:
            ws2.append([p.product_name, p.quantity_sold, p.revenue])

        # Top Categories sheet
        ws3 = wb.create_sheet("Top Categories")
        ws3.append(["Category", "Quantity Sold", "Revenue"])
        for c in dashboard.top_categories:
            ws3.append([c.category_name, c.quantity_sold, c.revenue])

        # Sales Trend sheet
        ws4 = wb.create_sheet("Sales Trend")
        ws4.append(["Date", "Orders", "Revenue"])
        for t in dashboard.sales_trend:
            ws4.append([t.date, t.orders, t.revenue])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=report.xlsx"},
        )

    elif format == "pdf":
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors

        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import os
        font_path = r"C:\Windows\Fonts\arial.ttf"
        bold_font_path = r"C:\Windows\Fonts\arialbd.ttf"
        base_font = "Helvetica"
        bold_font = "Helvetica-Bold"
        if os.path.exists(font_path) and os.path.exists(bold_font_path):
            try:
                pdfmetrics.registerFont(TTFont('Unicode', font_path))
                pdfmetrics.registerFont(TTFont('Unicode-Bold', bold_font_path))
                base_font = "Unicode"
                bold_font = "Unicode-Bold"
            except Exception:
                pass

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        # Override fonts in styles
        styles["Title"].fontName = bold_font
        styles["Heading2"].fontName = bold_font
        styles["Normal"].fontName = base_font
        elements = []

        elements.append(Paragraph("CafePOS Sales Report", styles["Title"]))
        elements.append(Spacer(1, 20))

        # Summary table
        summary_data = [
            ["Metric", "Value"],
            ["Total Orders", str(dashboard.total_orders)],
            ["Total Revenue", f"₹{dashboard.total_revenue:.2f}"],
            ["Average Order Value", f"₹{dashboard.average_order_value:.2f}"],
        ]
        t = Table(summary_data, colWidths=[200, 200])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#714B67")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("FONTNAME", (0, 0), (-1, 0), bold_font),
            ("FONTNAME", (0, 1), (-1, -1), base_font),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 20))

        # Top products
        if dashboard.top_products:
            elements.append(Paragraph("Top Products", styles["Heading2"]))
            prod_data = [["Product", "Qty", "Revenue"]]
            for p in dashboard.top_products:
                prod_data.append([p.product_name, str(p.quantity_sold), f"₹{p.revenue:.2f}"])
            t2 = Table(prod_data, colWidths=[200, 80, 120])
            t2.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#714B67")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), bold_font),
                ("FONTNAME", (0, 1), (-1, -1), base_font),
            ]))
            elements.append(t2)

        doc.build(elements)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=report.pdf"},
        )
